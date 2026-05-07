"""백그라운드 작업 처리 — threading + queue, Celery 불필요"""
import io
import sys
import threading
import queue
import time
import shutil
from pathlib import Path
from typing import Optional

import pi_review_tool
import pdf_report as _pdf_report
from config import JOBS_DIR, JOB_TTL_SECONDS

# ---------------------------------------------------------------------------
# 공유 상태
# ---------------------------------------------------------------------------
job_states: dict[str, dict] = {}
_state_lock = threading.Lock()
_work_queue: queue.Queue = queue.Queue()


def _update(job_id: str, progress: int, message: str) -> None:
    with _state_lock:
        if job_id in job_states:
            job_states[job_id]["progress"] = progress
            job_states[job_id]["message"] = message


def _schedule_cleanup(job_id: str) -> None:
    def _cleanup():
        job_dir = JOBS_DIR / job_id
        if job_dir.exists():
            shutil.rmtree(job_dir, ignore_errors=True)
        with _state_lock:
            job_states.pop(job_id, None)
    t = threading.Timer(JOB_TTL_SECONDS, _cleanup)
    t.daemon = True
    t.start()


# ---------------------------------------------------------------------------
# stdout 훅 — pi_review_tool.run() print 출력을 progress로 변환
# ---------------------------------------------------------------------------
class _ProgressWriter(io.TextIOBase):
    def __init__(self, job_id: str, orig: object) -> None:
        self.job_id = job_id
        self.orig = orig
        self._ocr_total = 0
        self._ocr_done = 0

    def write(self, text: str) -> int:
        try:
            self.orig.write(text)  # type: ignore[union-attr]
            self.orig.flush()      # type: ignore[union-attr]
        except Exception:
            pass
        t = text.strip()
        if not t:
            return len(text)
        if "PI PDF:" in t:
            try:
                self._ocr_total = int(t.split("PI PDF:")[1].strip().split("개")[0].strip())
            except Exception:
                pass
            _update(self.job_id, 20, f"PI PDF {self._ocr_total}개 발견")
        elif "박스내용:" in t:
            _update(self.job_id, 10, "박스내용 엑셀 읽는 중...")
        elif "CI/PL:" in t:
            _update(self.job_id, 15, "CI/PL 파일 확인 중...")
        elif "EasyOCR 초기화" in t:
            _update(self.job_id, 28, "OCR 엔진 초기화 중... (최초 실행 시 ~1분)")
        elif "[OCR]" in t:
            self._ocr_done += 1
            name = t.replace("[OCR]", "").strip().rstrip(".")
            if len(name) > 35:
                name = name[:32] + "..."
            total = max(self._ocr_total, 1)
            pct = 30 + int((self._ocr_done / total) * 45)
            _update(self.job_id, pct, f"OCR 처리 중: {name}")
        elif "OCR 완료:" in t:
            _update(self.job_id, 78, "OCR 완료, 비교 분석 중...")
        elif "CI 항목:" in t:
            _update(self.job_id, 80, "CI/PL 항목 비교 중...")
        return len(text)

    def flush(self) -> None:
        try:
            self.orig.flush()  # type: ignore[union-attr]
        except Exception:
            pass


# ---------------------------------------------------------------------------
# 워커 루프 (데몬 스레드 1개)
# ---------------------------------------------------------------------------
def _worker_loop() -> None:
    while True:
        job_id, zip_path, output_dir = _work_queue.get()
        orig_stdout = sys.stdout
        try:
            sys.stdout = _ProgressWriter(job_id, orig_stdout)
            _update(job_id, 5, "파일 압축 해제 중...")

            excel_path: Optional[Path] = pi_review_tool.run(zip_path, output_dir)

            _update(job_id, 88, "PDF 보고서 생성 중...")
            pdf_path = _pdf_report.generate(excel_path, output_dir)

            summary_data = _parse_summary(excel_path)
            _update(job_id, 100, "완료")
            with _state_lock:
                job_states[job_id]["status"] = "done"
                job_states[job_id]["excel_path"] = str(excel_path)
                job_states[job_id]["pdf_path"] = str(pdf_path)
                job_states[job_id]["summary_data"] = summary_data
            _schedule_cleanup(job_id)

        except Exception as exc:
            with _state_lock:
                job_states[job_id]["status"] = "error"
                job_states[job_id]["message"] = str(exc)
        finally:
            sys.stdout = orig_stdout
            _work_queue.task_done()


def _parse_summary(excel_path: Path) -> dict:
    """요약 시트에서 비교별 상태 카운트 추출"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        if "요약" not in wb.sheetnames:
            return {}
        ws = wb["요약"]
        result: dict = {}
        current: str | None = None
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0])
            val = row[1]
            stripped = key.strip()
            if stripped.startswith("[") and stripped.endswith("]"):
                current = stripped[1:-1]
                result[current] = {}
            elif current and key.startswith("  "):
                count = int(val) if isinstance(val, (int, float)) else 0
                result[current][stripped] = count
        return result
    except Exception:
        return {}


def start_worker() -> None:
    t = threading.Thread(target=_worker_loop, daemon=True, name="pi-worker")
    t.start()


def enqueue(job_id: str, zip_path: Path, output_dir: Path) -> None:
    with _state_lock:
        job_states[job_id] = {
            "status": "running",
            "progress": 0,
            "message": "대기 중...",
            "excel_path": None,
            "pdf_path": None,
        }
    _work_queue.put((job_id, zip_path, output_dir))
