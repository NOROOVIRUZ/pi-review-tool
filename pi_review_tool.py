#!/usr/bin/env python3
"""PI / CI / PL 3-way 검토 툴 (v2)

ZIP 또는 폴더 안의 문서를 PI 기준으로 비교합니다:
  - PI PDF   : 원본 계약서 (모든 비교의 기준)
  - 박스내용 xlsx : OA 내부 입고 계획 (PI 번호로 연결)
  - CI / PL  : 공급사 상업송장/패킹리스트 ('CI'/'PL' 시트 보유 xls/xlsx)

비교 결과를 new/ 폴더에 multi-sheet 엑셀로 출력합니다.
"""

from __future__ import annotations

import argparse
import json
import re
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

try:
    import xlrd as _xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import easyocr as _easyocr
    from pdf2image import convert_from_path as _pdf_to_images
    import numpy as _np
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

_OCR_READER = None


def _get_ocr_reader():
    global _OCR_READER
    if _OCR_READER is None:
        print("  [OCR] EasyOCR 초기화 중... (첫 실행시 모델 다운로드 ~100MB)", flush=True)
        import warnings, logging
        warnings.filterwarnings("ignore")
        logging.getLogger("easyocr").setLevel(logging.ERROR)
        _OCR_READER = _easyocr.Reader(["ko", "en"], gpu=False, verbose=False)
    return _OCR_READER

# ---------------------------------------------------------------------------
# 정규식 상수
# ---------------------------------------------------------------------------
HANGUL_RE = re.compile(r"[가-힣]")
ORDER_RE = re.compile(r"주문번호\s*[:：]\s*(\d{9})")
LOT_RE = re.compile(r"로트번호\s*[:：]\s*([^\n\r]+)")
PI_RE = re.compile(r"PI\s*(?:NO\.?|No\.?)\s*[:：]?\s*(\d{9})", re.IGNORECASE)
MODEL_RE = re.compile(r"[A-Z]{2,}[A-Z0-9-]{2,}")
# CI/PL 설명에서 괄호 안 모델코드 추출 (전각 괄호 포함)
PAREN_MODEL_RE = re.compile(r"[（(]([A-Z][A-Z0-9\-]{2,})[）)]")

BLOCKED_MODELS = {
    "FOC", "ADD", "NO", "SALE", "USD", "FOB", "NINGBO", "TOTAL",
    "PRICE", "PAID", "PHOTO", "QUANTITY", "BLUE", "BEIGE", "PINK",
    "GRAY", "GREY", "BLACK", "WHITE", "PCS", "CTN", "CTNS", "CLS",
    "VSL", "ETD", "ETA", "FCL", "LCL", "OA",
}

# ---------------------------------------------------------------------------
# OCR 텍스트 정제 & 모델코드 보정
# ---------------------------------------------------------------------------

def _fix_ocr_price_text(text: str) -> str:
    """OCR 오인식된 가격 표기를 정정합니다."""
    # 달러 기호: S 또는 $ 앞 공백 등 → $
    text = re.sub(r"\bS(\d)", r"$\1", text)
    # 유럽식 소수점 (5,35 → 5.35): 숫자,두자리숫자 패턴만 처리
    text = re.sub(r"(\b\d+),(\d{2})\b(?![\d,])", lambda m: f"{m.group(1)}.{m.group(2)}", text)
    return text


def _fix_ocr_model_digits(model: str) -> str:
    """모델코드 숫자 자리 OCR 오인식을 정정합니다 (예: OD7→007, OO→00)."""
    if "-" not in model:
        return model
    prefix, suffix = model.split("-", 1)
    fixed_suffix = ""
    for i, c in enumerate(suffix):
        if c.isdigit():
            fixed_suffix += c
        elif c in "Oo":
            fixed_suffix += "0"
        elif c in "Iil":
            # 알파벳 섹션(끝부분)에서는 그대로, 숫자 섹션에서는 1로
            is_digit_zone = any(ch.isdigit() for ch in suffix[:i])
            fixed_suffix += "1" if is_digit_zone else c
        elif c == "D" and i < len(suffix) - 2 and suffix[i + 1].isdigit():
            fixed_suffix += "0"
        elif c == "d" and i < len(suffix) - 2:
            fixed_suffix += "0"
        else:
            fixed_suffix += c
    return prefix + "-" + fixed_suffix


def _fuzzy_match_model(model: str, known_models: frozenset[str]) -> str | None:
    """OCR 오인식된 모델코드를 알려진 모델목록과 매칭합니다."""
    if not model or model in BLOCKED_MODELS:
        return None
    if model in known_models:
        return model

    # 1단계: 숫자 자리 치환 정정
    fixed = _fix_ocr_model_digits(model)
    if fixed in known_models:
        return fixed

    # 2단계: difflib 유사도 매칭
    from difflib import get_close_matches
    for cutoff in (0.80, 0.70, 0.65):
        for candidate in (model, fixed):
            matches = get_close_matches(candidate, known_models, n=1, cutoff=cutoff)
            if matches:
                return matches[0]
    return None


def normalize_ocr_models(pi_docs: list[PiDocument], known_models: frozenset[str]) -> None:
    """OCR로 파싱된 PI 문서의 모델코드를 알려진 모델로 보정합니다."""
    for doc in pi_docs:
        if doc.status != "PARSED_OCR":
            continue
        for item in doc.items:
            if item.model and item.model not in known_models:
                corrected = _fuzzy_match_model(item.model, known_models)
                if corrected and corrected != item.model:
                    item.evidence = f"[OCR보정:{item.model}→{corrected}] " + item.evidence
                    item.model = corrected


def ocr_pdf_pages(path: Path) -> list[str]:
    """EasyOCR로 PDF 각 페이지를 OCR하여 페이지별 텍스트 리스트를 반환합니다."""
    import warnings, io
    warnings.filterwarnings("ignore")
    reader = _get_ocr_reader()
    page_images = _pdf_to_images(str(path), dpi=200)
    page_texts: list[str] = []
    for page_img in page_images:
        img = _np.array(page_img)
        results = reader.readtext(img, detail=0, paragraph=False)
        lines = [ln.strip() for ln in results if len(ln.strip()) >= 2]
        raw = "\n".join(lines)
        page_texts.append(_fix_ocr_price_text(raw))
    return page_texts


# ---------------------------------------------------------------------------
# 데이터 클래스
# ---------------------------------------------------------------------------

@dataclass
class ExcelRow:
    row_no: int
    product: str
    description: str
    carton_unit: float | None
    carton_count: float | None
    quantity: float | None
    extra_stock: str
    note: str
    order_no: str | None
    lot_no: str | None
    model: str | None
    item_class: str


@dataclass
class CiItem:
    line_no: int
    description: str
    model: str | None
    volt: str
    qty: float | None
    unit: str
    unit_price: float | None
    amount: float | None
    item_class: str  # MAIN / FOC / BOX


@dataclass
class PlItem:
    line_no: int
    ctn_range: str
    description: str
    model: str | None
    ea_per_carton: float | None
    ctns: float | None
    qty: float | None
    nw: float | None
    gw: float | None
    meas: float | None
    item_class: str  # MAIN / FOC / BOX


@dataclass
class PiItem:
    pi_no: str
    pdf_name: str
    page: int
    item_class: str
    product: str
    model: str | None
    ea_per_carton: float | None
    quantity: float | None
    unit_price: float | None
    total_price: float | None
    evidence: str


@dataclass
class PiDocument:
    path: Path
    pi_no: str | None
    status: str
    page_count: int
    char_count: int
    lot_text: str | None = None
    items: list[PiItem] = field(default_factory=list)
    error: str | None = None
    text_preview: str = ""


@dataclass
class BoxCompareRow:
    """PI vs 박스내용 비교 결과"""
    status: str
    severity: str
    excel: ExcelRow | None
    pi: PiItem | None
    field: str
    excel_value: str
    pi_value: str
    diff: str
    request: str
    pdf_name: str
    evidence: str


@dataclass
class CiCompareRow:
    """PI vs CI 비교 결과"""
    status: str
    severity: str
    model: str | None
    description: str
    item_class: str
    pi_no: str | None
    pi_qty: float | None
    ci_qty: float | None
    qty_diff: str
    pi_unit_price: float | None
    ci_unit_price: float | None
    price_diff: str
    pi_total: float | None
    ci_amount: float | None
    amount_diff: str
    request: str
    box_qty: float | None
    pdf_name: str


@dataclass
class PlCompareRow:
    """PI vs PL 비교 결과"""
    status: str
    severity: str
    model: str | None
    description: str
    item_class: str
    pi_no: str | None
    pi_qty: float | None
    pl_qty: float | None
    qty_diff: str
    pi_ea_per_ctn: float | None
    pl_ea_per_ctn: float | None
    ea_diff: str
    pi_ctns_calc: float | None
    pl_ctns: float | None
    ctns_diff: str
    request: str
    box_qty: float | None
    pdf_name: str


# ---------------------------------------------------------------------------
# 유틸리티
# ---------------------------------------------------------------------------

def norm_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def to_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return None if v != v else v  # NaN guard
    text = norm_text(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def fmt_num(value: float | int | None) -> str:
    if value is None:
        return ""
    if abs(float(value) - round(float(value))) < 1e-9:
        return str(int(round(float(value))))
    return f"{float(value):,.2f}".rstrip("0").rstrip(".")


def extract_model(*parts: str) -> str | None:
    joined = " ".join(norm_text(part) for part in parts if part)
    candidates = MODEL_RE.findall(joined)
    filtered = [c for c in candidates if c not in BLOCKED_MODELS and not c.startswith("LOT")]
    return filtered[-1] if filtered else None


def find_model_in_ci_pl(description: str) -> str | None:
    """CI/PL 설명 문자열에서 괄호 안 모델코드를 추출합니다."""
    text = unicodedata.normalize("NFC", description).upper()
    text = text.replace("（", "(").replace("）", ")")
    for m in PAREN_MODEL_RE.finditer(text):
        candidate = m.group(1)
        if candidate not in BLOCKED_MODELS and len(candidate) >= 4:
            return candidate
    # fallback: 괄호 없이 직접 매칭
    return extract_model(description)


def classify_excel_row(description: str, product: str) -> str:
    text = f"{description} {product}".upper()
    if "FOC" in text or "무료" in text:
        if "BOX" in text or "박스" in text or "MAIL" in text or "COLOR" in text or "COLOUR" in text:
            return "BOX"
        return "FOC"
    if "BOX" in text or "박스" in text or "MAIL" in text or "COLOR" in text or "COLOUR" in text:
        return "BOX"
    return "MAIN"


def classify_ci_pl_item(description: str) -> str:
    text = norm_text(description).upper()
    if any(k in text for k in ["BOX", "MAIL", "COLOUR", "COLOR", "彩盒", "邮购盒"]):
        return "BOX"
    if "FOC" in text:
        return "FOC"
    return "MAIN"


# ---------------------------------------------------------------------------
# 입력 파일 탐색
# ---------------------------------------------------------------------------

def find_input_root(input_path: Path) -> tuple[Path, tempfile.TemporaryDirectory[str] | None]:
    input_path = input_path.expanduser().resolve()
    if input_path.is_dir():
        return input_path, None
    if not input_path.exists():
        raise FileNotFoundError(f"입력 경로가 없습니다: {input_path}")
    if input_path.suffix.lower() != ".zip":
        raise ValueError("입력은 ZIP 파일 또는 압축 해제된 폴더여야 합니다.")

    temp_dir = tempfile.TemporaryDirectory(prefix="pi_review_")
    root = Path(temp_dir.name)
    with zipfile.ZipFile(input_path) as zf:
        for info in zf.infolist():
            if "__MACOSX" in info.filename:
                continue
            target = root / Path(info.filename).name
            if not info.is_dir():
                target.write_bytes(zf.read(info.filename))
    return root, temp_dir


def choose_excel(root: Path) -> Path:
    candidates = sorted(root.rglob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("비교할 .xlsx 파일을 찾지 못했습니다.")
    preferred = [p for p in candidates if "박스내용" in unicodedata.normalize("NFC", p.name)]
    return preferred[0] if preferred else candidates[0]


def find_ci_pl_file(root: Path) -> Path | None:
    """CI/PL 시트가 있는 xls/xlsx 파일을 찾습니다."""
    for ext_pattern in ["*.xls", "*.xlsx"]:
        for p in sorted(root.rglob(ext_pattern)):
            if "__MACOSX" in str(p):
                continue
            if "박스내용" in unicodedata.normalize("NFC", p.name):
                continue
            try:
                if p.suffix.lower() == ".xls" and HAS_XLRD:
                    wb = _xlrd.open_workbook(str(p))
                    if "CI" in wb.sheet_names() or "PL" in wb.sheet_names():
                        return p
                elif p.suffix.lower() == ".xlsx":
                    wb = load_workbook(p, read_only=True)
                    if "CI" in wb.sheetnames or "PL" in wb.sheetnames:
                        wb.close()
                        return p
            except Exception:
                pass
    return None


def find_pdfs(root: Path) -> list[Path]:
    return sorted(root.rglob("*.pdf"))


# ---------------------------------------------------------------------------
# 박스내용 파싱 (기존)
# ---------------------------------------------------------------------------

def read_excel_rows(excel_path: Path) -> list[ExcelRow]:
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook.active

    header_row = None
    headers: dict[str, int] = {}
    for row in worksheet.iter_rows():
        values = [norm_text(cell.value) for cell in row]
        if "제품" in values and "수량" in values:
            header_row = row[0].row
            headers = {value: idx + 1 for idx, value in enumerate(values) if value}
            break
    if header_row is None:
        raise ValueError("엑셀에서 '제품/수량' 헤더 행을 찾지 못했습니다.")

    def cell(row_no: int, name: str) -> object:
        col = headers.get(name)
        return worksheet.cell(row_no, col).value if col else None

    current_order: str | None = None
    current_lot: str | None = None
    model_order: dict[str, str | None] = {}
    model_lot: dict[str, str | None] = {}
    rows: list[ExcelRow] = []

    for row_no in range(header_row + 1, worksheet.max_row + 1):
        product = norm_text(cell(row_no, "제품"))
        description = norm_text(cell(row_no, "구분"))
        note = norm_text(cell(row_no, "비고"))
        quantity = to_number(cell(row_no, "수량"))

        if not product and not description:
            continue
        if product.startswith("총") or description.startswith("총"):
            continue
        if product == "제품" or description == "구분":
            continue

        order_match = ORDER_RE.search(note)
        lot_match = LOT_RE.search(note)
        if order_match:
            current_order = order_match.group(1)
        if lot_match:
            current_lot = lot_match.group(1).strip()

        item_class = classify_excel_row(description, product)
        model = extract_model(product, description)
        row_order = current_order
        row_lot = current_lot
        if not order_match and model in model_order:
            row_order = model_order[model]
            row_lot = model_lot.get(model)

        if order_match and model:
            model_order[model] = current_order
            model_lot[model] = current_lot

        rows.append(
            ExcelRow(
                row_no=row_no,
                product=product,
                description=description,
                carton_unit=to_number(cell(row_no, "입수단위")),
                carton_count=to_number(cell(row_no, "카톤수")),
                quantity=quantity,
                extra_stock=norm_text(cell(row_no, "재고 따로 추가")),
                note=note,
                order_no=row_order,
                lot_no=row_lot,
                model=model,
                item_class=item_class,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# CI / PL 파싱
# ---------------------------------------------------------------------------

def _rows_from_xlrd(sheet) -> list[list]:
    return [sheet.row_values(i) for i in range(sheet.nrows)]


def _parse_ci_rows(rows: list[list]) -> list[CiItem]:
    """CI 시트 행 목록에서 CiItem 리스트를 반환합니다."""
    items: list[CiItem] = []
    header_found = False
    line_no = 0

    for row in rows:
        if not header_found:
            row_str = " ".join(str(v) for v in row)
            if "Quantity/Unit" in row_str or "14)" in row_str:
                header_found = True
            continue

        # description: col[5], qty: col[7]
        raw_desc = row[5] if len(row) > 5 else None
        raw_qty = row[7] if len(row) > 7 else None
        raw_item_no = row[3] if len(row) > 3 else None

        description = norm_text(raw_desc)
        qty = to_number(raw_qty)
        item_no_str = norm_text(raw_item_no)

        # 총계 행에서 멈춤
        if "TOTAL" in item_no_str.upper() or "TOTAL" in description.upper():
            break

        if not description or qty is None or qty <= 0:
            continue

        # item_no가 숫자인 행만 처리
        try:
            float(item_no_str)
        except (ValueError, TypeError):
            continue

        volt = norm_text(row[6] if len(row) > 6 else None)
        unit = norm_text(row[8] if len(row) > 8 else None) or "PCS"
        unit_price = to_number(row[9] if len(row) > 9 else None)
        amount = to_number(row[10] if len(row) > 10 else None)

        line_no += 1
        items.append(CiItem(
            line_no=line_no,
            description=description,
            model=find_model_in_ci_pl(description),
            volt=volt,
            qty=qty,
            unit=unit,
            unit_price=unit_price,
            amount=amount,
            item_class=classify_ci_pl_item(description),
        ))

    return items


def _parse_pl_rows(rows: list[list]) -> list[PlItem]:
    """PL 시트 행 목록에서 PlItem 리스트를 반환합니다."""
    items: list[PlItem] = []
    header_found = False
    skip_units_row = False
    line_no = 0

    for row in rows:
        if not header_found:
            row_str = " ".join(str(v) for v in row)
            if "DESCRIPTIONS" in row_str and "PACKING" in row_str:
                header_found = True
                skip_units_row = True
            continue

        if skip_units_row:
            skip_units_row = False
            continue

        raw_ctn = row[1] if len(row) > 1 else None
        raw_desc = row[2] if len(row) > 2 else None
        raw_qty = row[5] if len(row) > 5 else None

        ctn_range = norm_text(raw_ctn)
        description = norm_text(raw_desc)
        qty = to_number(raw_qty)

        # 총계 행 감지
        if "total" in ctn_range.lower() or "total" in description.lower():
            break

        if not description or qty is None or qty <= 0:
            continue

        ea_per_ctn = to_number(row[3] if len(row) > 3 else None)
        ctns = to_number(row[4] if len(row) > 4 else None)
        nw = to_number(row[6] if len(row) > 6 else None)
        gw = to_number(row[7] if len(row) > 7 else None)
        meas = to_number(row[8] if len(row) > 8 else None)

        line_no += 1
        items.append(PlItem(
            line_no=line_no,
            ctn_range=ctn_range,
            description=description,
            model=find_model_in_ci_pl(description),
            ea_per_carton=ea_per_ctn,
            ctns=ctns,
            qty=qty,
            nw=nw,
            gw=gw,
            meas=meas,
            item_class=classify_ci_pl_item(description),
        ))

    return items


def parse_ci_pl(path: Path) -> tuple[list[CiItem], list[PlItem]]:
    """xls/xlsx 파일에서 CI, PL 항목을 파싱합니다."""
    ci_items: list[CiItem] = []
    pl_items: list[PlItem] = []

    if path.suffix.lower() == ".xls":
        if not HAS_XLRD:
            print(f"  [경고] xlrd 미설치로 {path.name} 파싱 불가")
            return [], []
        wb = _xlrd.open_workbook(str(path))
        sheet_names = wb.sheet_names()
        if "CI" in sheet_names:
            ci_items = _parse_ci_rows(_rows_from_xlrd(wb.sheet_by_name("CI")))
        if "PL" in sheet_names:
            pl_items = _parse_pl_rows(_rows_from_xlrd(wb.sheet_by_name("PL")))

    elif path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        if "CI" in wb.sheetnames:
            ws = wb["CI"]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            ci_items = _parse_ci_rows(rows)
        if "PL" in wb.sheetnames:
            ws = wb["PL"]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            pl_items = _parse_pl_rows(rows)
        wb.close()

    return ci_items, pl_items


# ---------------------------------------------------------------------------
# PI PDF 파싱 (개선됨)
# ---------------------------------------------------------------------------

def safe_pdf_text(page) -> str:
    try:
        return page.extract_text(extraction_mode="layout") or ""
    except TypeError:
        return page.extract_text() or ""


def extract_lot_text(text: str) -> str | None:
    for line in text.splitlines():
        if "LOT" in line.upper() and "NO" in line.upper():
            return norm_text(line)[:300]
    return None


def number_from_money(value: str | None) -> float | None:
    if not value:
        return None
    text = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def split_supplier_blocks(page_text: str) -> list[str]:
    chunks = re.split(r"(?=Supplier Model Name)", page_text)
    return [chunk for chunk in chunks if "Buyer Model Name" in chunk and "$" in chunk]


def clean_product_candidate(line: str) -> str:
    text = norm_text(line)
    text = re.sub(r"^Supplier Model Name\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^Buyer Model Name\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^[A-Z0-9-]{3,}\s+", "", text)
    text = re.sub(r"\s+\d+\s+\$.*$", "", text)
    return text.strip()


def parse_main_items(pi_no: str, pdf_name: str, page_text: str) -> list[PiItem]:
    items: list[PiItem] = []
    for block in split_supplier_blocks(page_text):
        compact = norm_text(block)
        model = None
        model_region = re.search(
            r"Buyer Model Name\s+(.*?)(?:Supplier Model Name|additional parts cost|Total Amount|Remarks|$)",
            compact,
            re.IGNORECASE,
        )
        if model_region:
            model = extract_model(model_region.group(1))
        if not model:
            model = extract_model(compact)

        qty_match = re.search(
            r"(?P<ea>\d+(?:\.\d+)?)\s+\$?\s*(?P<unit>\d[\d,]*(?:\.\d+)?)\s+"
            r"(?P<qty>\d[\d,]*)\s+\$?\s*(?P<total>\d[\d,]*(?:\.\d+)?)",
            compact,
        )
        if not qty_match:
            continue

        product = ""
        for raw_line in block.splitlines():
            line = clean_product_candidate(raw_line)
            if HANGUL_RE.search(line) and "Remarks" not in line and "Product Name" not in line:
                product = line
                break
        if not product:
            product = model or "PI item"

        items.append(PiItem(
            pi_no=pi_no,
            pdf_name=pdf_name,
            page=1,
            item_class="MAIN",
            product=product,
            model=model,
            ea_per_carton=number_from_money(qty_match.group("ea")),
            quantity=number_from_money(qty_match.group("qty")),
            unit_price=number_from_money(qty_match.group("unit")),
            total_price=number_from_money(qty_match.group("total")),
            evidence=compact[:500],
        ))
    return items


def appendix_model_candidates(page_text: str) -> list[str]:
    models = []
    for line in page_text.splitlines():
        for model in MODEL_RE.findall(line):
            if model not in BLOCKED_MODELS:
                models.append(model)
    return models


def parse_appendix_items(
    pi_no: str,
    pdf_name: str,
    page_no: int,
    page_text: str,
    main_items: list[PiItem],
) -> list[PiItem]:
    items: list[PiItem] = []
    lines = [line for line in page_text.splitlines() if norm_text(line)]
    main_models = [item.model for item in main_items if item.model]
    appendix_models = appendix_model_candidates(page_text)
    previous_context: list[str] = []

    for raw_line in lines:
        line = norm_text(raw_line)
        if "$0.00" not in line:
            if not any(skip in line for skip in ["CLIENT INFO", "SUPPLIER INFO", "Company:", "Email:", "Tel:"]):
                previous_context.append(line)
                previous_context = previous_context[-4:]
            continue

        numbers = re.findall(r"(?<![\d.])\d{1,5}(?![\d.])", line)
        numbers = [n for n in numbers if n not in {pi_no, "0"}]
        if not numbers:
            continue
        quantity = float(numbers[-1])
        context = " ".join(previous_context + [line])
        line_upper = line.upper()
        has_foc = any(w in context.upper() for w in ["FOC", "무료", "免费", "本품", "本品"])
        has_box = any(w in context.upper() for w in ["BOX", "MAIL", "彩盒", "邮购盒", "박스", "COLOR", "COLOUR"])
        line_has_box = any(w in line_upper for w in ["BOX", "MAIL", "彩盒", "邮购盒", "박스", "COLOR", "COLOUR"])

        if "Total Amount" in context and not has_foc and not has_box:
            continue

        if line_has_box or has_box:
            item_class = "BOX"
        elif has_foc:
            item_class = "FOC"
        else:
            item_class = "BOX"

        model = extract_model(context)
        if not model:
            for candidate in appendix_models:
                if candidate in context:
                    model = candidate
                    break
        if not model and len(main_models) == 1:
            model = main_models[0]

        product_candidates = [c for c in previous_context if HANGUL_RE.search(c) or "BOX" in c.upper()]
        product = product_candidates[-1] if product_candidates else context[:80]
        items.append(PiItem(
            pi_no=pi_no,
            pdf_name=pdf_name,
            page=page_no,
            item_class=item_class,
            product=product,
            model=model,
            ea_per_carton=None,
            quantity=quantity,
            unit_price=0.0,
            total_price=0.0,
            evidence=context[:500],
        ))
        previous_context = []
    return dedupe_pi_items(items)


def dedupe_pi_items(items: list[PiItem]) -> list[PiItem]:
    seen: set[tuple[str, str | None, float | None, str]] = set()
    result: list[PiItem] = []
    for item in items:
        key = (item.item_class, item.model, item.quantity, item.product)
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def read_pdf(path: Path) -> PiDocument:
    fallback_pi = re.match(r"^(\d{9})", path.name)
    try:
        reader = PdfReader(str(path))
        pages = [safe_pdf_text(page) for page in reader.pages]
        full_text = "\n\n".join(pages)
        char_count = len(full_text.strip())
        pi_match = PI_RE.search(full_text)
        pi_no = pi_match.group(1) if pi_match else (fallback_pi.group(1) if fallback_pi else None)
        lot_text = extract_lot_text(full_text)

        if char_count < 50:
            if not HAS_OCR:
                return PiDocument(
                    path=path,
                    pi_no=pi_no,
                    status="OCR_REQUIRED",
                    page_count=len(reader.pages),
                    char_count=char_count,
                    lot_text=lot_text,
                    error="스캔 PDF입니다. easyocr+pdf2image 설치 시 자동 처리됩니다.",
                    text_preview=full_text[:500],
                )
            # OCR 처리
            print(f"  [OCR] {path.name} ...", flush=True)
            try:
                ocr_pages = ocr_pdf_pages(path)
            except Exception as ocr_exc:
                return PiDocument(
                    path=path,
                    pi_no=pi_no,
                    status="OCR_FAILED",
                    page_count=len(reader.pages),
                    char_count=0,
                    error=f"OCR 실패: {ocr_exc}",
                )
            ocr_full = "\n\n".join(ocr_pages)
            ocr_char_count = len(ocr_full.strip())
            pi_match2 = PI_RE.search(ocr_full)
            pi_no = pi_match2.group(1) if pi_match2 else pi_no
            lot_text2 = extract_lot_text(ocr_full)
            doc = PiDocument(
                path=path,
                pi_no=pi_no,
                status="PARSED_OCR",
                page_count=len(reader.pages),
                char_count=ocr_char_count,
                lot_text=lot_text2,
                text_preview=ocr_full[:500],
            )
            if pi_no:
                doc.items.extend(parse_main_items(pi_no, path.name, ocr_pages[0] if ocr_pages else ""))
                for page_idx, page_text in enumerate(ocr_pages[1:], start=2):
                    doc.items.extend(parse_appendix_items(pi_no, path.name, page_idx, page_text, doc.items))
            return doc

        doc = PiDocument(
            path=path,
            pi_no=pi_no,
            status="PARSED",
            page_count=len(reader.pages),
            char_count=char_count,
            lot_text=lot_text,
            text_preview=full_text[:500],
        )
        if pi_no:
            doc.items.extend(parse_main_items(pi_no, path.name, pages[0] if pages else ""))
            for page_idx, page_text in enumerate(pages[1:], start=2):
                doc.items.extend(parse_appendix_items(pi_no, path.name, page_idx, page_text, doc.items))
        return doc
    except Exception as exc:
        return PiDocument(
            path=path,
            pi_no=fallback_pi.group(1) if fallback_pi else None,
            status="ERROR",
            page_count=0,
            char_count=0,
            error=str(exc),
        )


# ---------------------------------------------------------------------------
# PI 인덱스 빌더
# ---------------------------------------------------------------------------

def build_pi_index(docs: list[PiDocument]) -> dict[tuple[str, str | None, str], list[PiItem]]:
    """(pi_no, model, item_class) -> [PiItem] 인덱스"""
    index: dict[tuple[str, str | None, str], list[PiItem]] = {}
    for doc in docs:
        for item in doc.items:
            key = (item.pi_no, item.model, item.item_class)
            index.setdefault(key, []).append(item)
    return index


def build_pi_model_index(docs: list[PiDocument]) -> dict[tuple[str, str], PiItem]:
    """(model, item_class) -> PiItem 인덱스 (모델코드로 CI/PL 매칭용)"""
    index: dict[tuple[str, str], PiItem] = {}
    for doc in docs:
        for item in doc.items:
            if item.model:
                key = (item.model, item.item_class)
                if key not in index:
                    index[key] = item
    return index


def find_doc_name(pi_docs: list[PiDocument], pi_no: str) -> str:
    for doc in pi_docs:
        if doc.pi_no == pi_no:
            return doc.path.name
    return ""


def find_doc_name_by_model(pi_docs: list[PiDocument], model: str) -> str:
    for doc in pi_docs:
        for item in doc.items:
            if item.model == model:
                return doc.path.name
    return ""


# ---------------------------------------------------------------------------
# PI vs 박스내용 비교 (기존 로직)
# ---------------------------------------------------------------------------

def _excel_group_key(excel: ExcelRow) -> tuple[str | None, str | None, str]:
    return (excel.order_no, excel.model, excel.item_class)


_FOC_ALIASES = frozenset({"FOC", "BOX"})


def _pick_pi_item(
    excel: ExcelRow,
    pi_index: dict[tuple[str, str | None, str], list[PiItem]],
    preferred_quantity: float | None = None,
) -> PiItem | None:
    if not excel.order_no:
        return None
    exact = pi_index.get((excel.order_no, excel.model, excel.item_class), [])
    if not exact and excel.item_class in _FOC_ALIASES:
        for alias in _FOC_ALIASES:
            exact = pi_index.get((excel.order_no, excel.model, alias), [])
            if exact:
                break
    if len(exact) == 1:
        return exact[0]
    if exact:
        if preferred_quantity is None:
            return exact[0]
        return min(exact, key=lambda item: abs((item.quantity or 0) - preferred_quantity))

    # 폴백: 모델코드 없는 행 전용 (모델이 명시된 경우 다른 모델로 대체 금지)
    if excel.model is not None:
        return None
    class_matches: list[PiItem] = []
    target_classes = _FOC_ALIASES if excel.item_class in _FOC_ALIASES else {excel.item_class}
    for (pi_no, _model, item_class), items in pi_index.items():
        if pi_no == excel.order_no and item_class in target_classes:
            class_matches.extend(items)
    if len(class_matches) == 1:
        return class_matches[0]
    if class_matches and preferred_quantity is not None:
        return min(class_matches, key=lambda item: abs((item.quantity or 0) - preferred_quantity))
    return None


def compare_pi_vs_box(excel_rows: list[ExcelRow], pi_docs: list[PiDocument]) -> list[BoxCompareRow]:
    pi_index = build_pi_index(pi_docs)
    ocr_pi_numbers = {doc.pi_no for doc in pi_docs if doc.status == "OCR_REQUIRED" and doc.pi_no}
    parsed_pi_numbers = {doc.pi_no for doc in pi_docs if doc.status in ("PARSED", "PARSED_OCR") and doc.pi_no}
    results: list[BoxCompareRow] = []

    group_quantities: dict[tuple, float] = {}
    group_counts: dict[tuple, int] = {}
    for row in excel_rows:
        key = _excel_group_key(row)
        group_counts[key] = group_counts.get(key, 0) + 1
        if row.quantity is not None:
            group_quantities[key] = group_quantities.get(key, 0.0) + row.quantity

    for excel in excel_rows:
        group_key = _excel_group_key(excel)
        group_count = group_counts.get(group_key, 1)
        excel_compare_qty = group_quantities.get(group_key, excel.quantity or 0.0)

        if not excel.order_no:
            results.append(BoxCompareRow(
                status="확인필요", severity="중간", excel=excel, pi=None,
                field="주문번호", excel_value="", pi_value="",
                diff="엑셀 행에 주문번호가 연결되지 않았습니다.",
                request="해당 행이 어느 PI 주문번호에 속하는지 확인해 주세요.",
                pdf_name="", evidence=excel.note,
            ))
            continue

        if excel.order_no in ocr_pi_numbers:
            results.append(BoxCompareRow(
                status="OCR필요", severity="중간", excel=excel, pi=None,
                field="PI PDF", excel_value=fmt_num(excel.quantity), pi_value="",
                diff="PI가 스캔 PDF라 텍스트 추출 비교를 완료하지 못했습니다.",
                request=f"주문번호 {excel.order_no} PI에 OCR을 적용한 뒤 수량/입수단위를 재검토해 주세요.",
                pdf_name=find_doc_name(pi_docs, excel.order_no), evidence="",
            ))
            continue

        if excel.order_no not in parsed_pi_numbers:
            results.append(BoxCompareRow(
                status="PI누락", severity="높음", excel=excel, pi=None,
                field="PI", excel_value=excel.order_no, pi_value="",
                diff="엑셀 주문번호에 해당하는 PI PDF를 찾지 못했습니다.",
                request=f"주문번호 {excel.order_no}의 PI PDF를 추가하거나 파일명을 확인해 주세요.",
                pdf_name="", evidence=excel.note,
            ))
            continue

        pi_item = _pick_pi_item(excel, pi_index, preferred_quantity=excel_compare_qty)
        if not pi_item:
            results.append(BoxCompareRow(
                status="PI항목누락", severity="높음", excel=excel, pi=None,
                field="제품/모델",
                excel_value=f"{excel.model or ''} {excel.product}".strip(),
                pi_value="",
                diff="PI에서 엑셀 행과 매칭되는 모델/항목을 찾지 못했습니다.",
                request=f"주문번호 {excel.order_no}에서 {excel.product} 항목이 PI에 있는지 확인해 주세요.",
                pdf_name=find_doc_name(pi_docs, excel.order_no),
                evidence=excel.description,
            ))
            continue

        row_status = "정상"
        diffs: list[str] = []
        requests: list[str] = []
        severity = "낮음"

        if pi_item.quantity is not None and abs(excel_compare_qty - pi_item.quantity) > 1e-9:
            row_status = "불일치"
            severity = "높음"
            label = "엑셀 합계" if group_count > 1 else "엑셀"
            diffs.append(f"수량: {label} {fmt_num(excel_compare_qty)} / PI {fmt_num(pi_item.quantity)}")
            requests.append(f"수량을 PI 기준 {fmt_num(pi_item.quantity)}로 맞출지 확인해 주세요.")

        if (
            excel.item_class == "MAIN"
            and group_count == 1
            and excel.carton_unit is not None
            and pi_item.ea_per_carton is not None
            and abs(excel.carton_unit - pi_item.ea_per_carton) > 1e-9
        ):
            row_status = "불일치"
            severity = "높음"
            diffs.append(f"입수단위: 엑셀 {fmt_num(excel.carton_unit)} / PI {fmt_num(pi_item.ea_per_carton)}")
            requests.append(f"입수단위를 PI 기준 {fmt_num(pi_item.ea_per_carton)}로 맞출지 확인해 주세요.")

        if not diffs:
            if group_count > 1:
                diffs.append(f"분할 포장 합계 일치. 엑셀 합계 {fmt_num(excel_compare_qty)} / PI {fmt_num(pi_item.quantity)}")
            else:
                diffs.append("비교 기준 값이 일치합니다.")
            requests.append("")

        results.append(BoxCompareRow(
            status=row_status,
            severity=severity,
            excel=excel,
            pi=pi_item,
            field="수량/입수단위",
            excel_value=f"행 수량 {fmt_num(excel.quantity)} / 합계 {fmt_num(excel_compare_qty)} / 입수 {fmt_num(excel.carton_unit)}",
            pi_value=f"수량 {fmt_num(pi_item.quantity)} / 입수 {fmt_num(pi_item.ea_per_carton)}",
            diff="; ".join(diffs),
            request=" ".join(r for r in requests if r),
            pdf_name=pi_item.pdf_name,
            evidence=pi_item.evidence,
        ))

    return results


# ---------------------------------------------------------------------------
# PI vs CI 비교
# ---------------------------------------------------------------------------

def compare_pi_vs_ci(
    pi_docs: list[PiDocument],
    ci_items: list[CiItem],
    box_rows: list[ExcelRow],
) -> list[CiCompareRow]:
    pi_model_index = build_pi_model_index(pi_docs)
    ocr_pi_nos = {doc.pi_no for doc in pi_docs if doc.status == "OCR_REQUIRED" and doc.pi_no}

    # 박스내용에서 model → {qty, pi_no} 맵
    box_qty_map: dict[str, float] = {}
    box_pi_map: dict[str, str] = {}
    for row in box_rows:
        if row.model and row.quantity:
            box_qty_map[row.model] = box_qty_map.get(row.model, 0.0) + row.quantity
        if row.model and row.order_no:
            box_pi_map[row.model] = row.order_no

    results: list[CiCompareRow] = []

    for ci in ci_items:
        if not ci.model:
            results.append(CiCompareRow(
                status="모델불명", severity="중간",
                model=None, description=ci.description, item_class=ci.item_class,
                pi_no=None, pi_qty=None, ci_qty=ci.qty, qty_diff="CI 항목에서 모델코드 미추출",
                pi_unit_price=None, ci_unit_price=ci.unit_price, price_diff="",
                pi_total=None, ci_amount=ci.amount, amount_diff="",
                request="CI 설명에서 모델코드를 직접 확인해 주세요.",
                box_qty=None, pdf_name="",
            ))
            continue

        pi_item = pi_model_index.get((ci.model, ci.item_class))
        if not pi_item and ci.item_class in _FOC_ALIASES:
            for alias in _FOC_ALIASES:
                pi_item = pi_model_index.get((ci.model, alias))
                if pi_item:
                    break

        # OCR 필요 PI에 속하는 모델
        box_pi_no = box_pi_map.get(ci.model)
        if not pi_item and box_pi_no in ocr_pi_nos:
            results.append(CiCompareRow(
                status="OCR필요", severity="중간",
                model=ci.model, description=ci.description, item_class=ci.item_class,
                pi_no=box_pi_no, pi_qty=None, ci_qty=ci.qty,
                qty_diff="PI가 스캔 PDF라 비교 불가",
                pi_unit_price=None, ci_unit_price=ci.unit_price, price_diff="",
                pi_total=None, ci_amount=ci.amount, amount_diff="",
                request=f"주문번호 {box_pi_no} PI OCR 후 재검토 필요.",
                box_qty=box_qty_map.get(ci.model), pdf_name="",
            ))
            continue

        if not pi_item:
            results.append(CiCompareRow(
                status="PI미매칭", severity="높음",
                model=ci.model, description=ci.description, item_class=ci.item_class,
                pi_no=None, pi_qty=None, ci_qty=ci.qty,
                qty_diff=f"CI의 {ci.model}({ci.item_class})을 PI에서 찾지 못했습니다.",
                pi_unit_price=None, ci_unit_price=ci.unit_price, price_diff="",
                pi_total=None, ci_amount=ci.amount, amount_diff="",
                request="PI에서 해당 모델/구분 항목을 확인해 주세요.",
                box_qty=box_qty_map.get(ci.model), pdf_name="",
            ))
            continue

        status = "정상"
        severity = "낮음"
        qty_diff = price_diff = amount_diff = ""
        requests: list[str] = []

        # ── 수량 비교 ──────────────────────────
        if pi_item.quantity is not None and ci.qty is not None:
            if abs(pi_item.quantity - ci.qty) > 1e-9:
                box_qty = box_qty_map.get(ci.model)
                if box_qty and abs(box_qty - ci.qty) < 1e-9:
                    status = "부분선적"
                    severity = "확인"
                    qty_diff = (
                        f"CI {fmt_num(ci.qty)} / PI전체 {fmt_num(pi_item.quantity)}"
                        f" → 박스내용 {fmt_num(box_qty)} 일치 (부분선적 가능)"
                    )
                    requests.append("의도적 부분선적이면 OK. 아니면 PI 수정 필요.")
                else:
                    status = "불일치"
                    severity = "높음"
                    qty_diff = f"CI {fmt_num(ci.qty)} / PI {fmt_num(pi_item.quantity)}"
                    requests.append(f"수량 PI 기준 {fmt_num(pi_item.quantity)}으로 수정 또는 부분선적 확인.")
            else:
                qty_diff = f"일치 ({fmt_num(ci.qty)})"

        # ── 단가 비교 (MAIN만) ─────────────────
        if ci.item_class == "MAIN" and pi_item.unit_price is not None and ci.unit_price is not None:
            if abs(pi_item.unit_price - ci.unit_price) > 0.001:
                status = "불일치"
                severity = "높음"
                price_diff = f"CI ${fmt_num(ci.unit_price)} / PI ${fmt_num(pi_item.unit_price)}"
                requests.append(f"단가 PI 기준 ${fmt_num(pi_item.unit_price)}로 수정 필요.")
            else:
                price_diff = f"일치 (${fmt_num(ci.unit_price)})"

        # ── 금액 검증: qty × unit_price (MAIN만) ─
        if ci.item_class == "MAIN" and ci.qty is not None and pi_item.unit_price is not None and ci.amount is not None:
            expected = (ci.qty or 0) * (pi_item.unit_price or 0)
            if abs(expected - ci.amount) > 1.0:
                status = "불일치"
                severity = "높음"
                amount_diff = f"CI ${fmt_num(ci.amount)} / 예상(qty×PI단가) ${fmt_num(expected)}"
                requests.append(f"CI 금액 ${fmt_num(expected)}으로 수정 필요.")
            else:
                amount_diff = f"일치 (${fmt_num(ci.amount)})"

        results.append(CiCompareRow(
            status=status, severity=severity,
            model=ci.model, description=ci.description, item_class=ci.item_class,
            pi_no=pi_item.pi_no, pi_qty=pi_item.quantity, ci_qty=ci.qty,
            qty_diff=qty_diff, pi_unit_price=pi_item.unit_price, ci_unit_price=ci.unit_price,
            price_diff=price_diff, pi_total=pi_item.total_price, ci_amount=ci.amount,
            amount_diff=amount_diff, request=" ".join(requests),
            box_qty=box_qty_map.get(ci.model),
            pdf_name=find_doc_name_by_model(pi_docs, ci.model),
        ))

    return results


# ---------------------------------------------------------------------------
# PI vs PL 비교
# ---------------------------------------------------------------------------

def compare_pi_vs_pl(
    pi_docs: list[PiDocument],
    pl_items: list[PlItem],
    box_rows: list[ExcelRow],
) -> list[PlCompareRow]:
    pi_model_index = build_pi_model_index(pi_docs)
    ocr_pi_nos = {doc.pi_no for doc in pi_docs if doc.status == "OCR_REQUIRED" and doc.pi_no}

    box_qty_map: dict[str, float] = {}
    box_pi_map: dict[str, str] = {}
    for row in box_rows:
        if row.model and row.quantity:
            box_qty_map[row.model] = box_qty_map.get(row.model, 0.0) + row.quantity
        if row.model and row.order_no:
            box_pi_map[row.model] = row.order_no

    # PL 항목을 (model, item_class)로 그룹화
    pl_groups: dict[tuple[str | None, str], list[PlItem]] = {}
    for item in pl_items:
        key = (item.model, item.item_class)
        pl_groups.setdefault(key, []).append(item)

    results: list[PlCompareRow] = []

    for (model, item_class), pl_group in pl_groups.items():
        pl_total_qty = sum(item.qty or 0 for item in pl_group)
        pl_total_ctns = sum(item.ctns or 0 for item in pl_group)
        # 주 입수단위: 카톤수 가장 많은 행 기준
        pl_primary = max(pl_group, key=lambda x: x.ctns or 0)
        pl_ea = pl_primary.ea_per_carton
        description = pl_group[0].description

        if not model:
            results.append(PlCompareRow(
                status="모델불명", severity="중간",
                model=None, description=description, item_class=item_class,
                pi_no=None, pi_qty=None, pl_qty=pl_total_qty,
                qty_diff="PL 항목에서 모델코드 미추출",
                pi_ea_per_ctn=None, pl_ea_per_ctn=pl_ea, ea_diff="",
                pi_ctns_calc=None, pl_ctns=pl_total_ctns, ctns_diff="",
                request="PL 설명에서 모델코드를 직접 확인해 주세요.",
                box_qty=None, pdf_name="",
            ))
            continue

        pi_item = pi_model_index.get((model, item_class))
        if not pi_item and item_class in _FOC_ALIASES:
            for alias in _FOC_ALIASES:
                pi_item = pi_model_index.get((model, alias))
                if pi_item:
                    break

        box_pi_no = box_pi_map.get(model)
        if not pi_item and box_pi_no in ocr_pi_nos:
            results.append(PlCompareRow(
                status="OCR필요", severity="중간",
                model=model, description=description, item_class=item_class,
                pi_no=box_pi_no, pi_qty=None, pl_qty=pl_total_qty,
                qty_diff="PI가 스캔 PDF라 비교 불가",
                pi_ea_per_ctn=None, pl_ea_per_ctn=pl_ea, ea_diff="",
                pi_ctns_calc=None, pl_ctns=pl_total_ctns, ctns_diff="",
                request=f"주문번호 {box_pi_no} PI OCR 후 재검토 필요.",
                box_qty=box_qty_map.get(model), pdf_name="",
            ))
            continue

        if not pi_item:
            results.append(PlCompareRow(
                status="PI미매칭", severity="높음",
                model=model, description=description, item_class=item_class,
                pi_no=None, pi_qty=None, pl_qty=pl_total_qty,
                qty_diff=f"PL의 {model}({item_class})을 PI에서 찾지 못했습니다.",
                pi_ea_per_ctn=None, pl_ea_per_ctn=pl_ea, ea_diff="",
                pi_ctns_calc=None, pl_ctns=pl_total_ctns, ctns_diff="",
                request="PI에서 해당 모델/구분 항목을 확인해 주세요.",
                box_qty=box_qty_map.get(model), pdf_name="",
            ))
            continue

        status = "정상"
        severity = "낮음"
        qty_diff = ea_diff = ctns_diff = ""
        requests: list[str] = []

        # ── 수량 비교 ──────────────────────────
        if pi_item.quantity is not None:
            box_qty = box_qty_map.get(model)
            if abs(pi_item.quantity - pl_total_qty) > 1e-9:
                if box_qty and abs(box_qty - pl_total_qty) < 1e-9:
                    status = "부분선적"
                    severity = "확인"
                    qty_diff = (
                        f"PL {fmt_num(pl_total_qty)} / PI전체 {fmt_num(pi_item.quantity)}"
                        f" → 박스내용 {fmt_num(box_qty)} 일치 (부분선적 가능)"
                    )
                    requests.append("의도적 부분선적이면 OK.")
                else:
                    status = "불일치"
                    severity = "높음"
                    qty_diff = f"PL {fmt_num(pl_total_qty)} / PI {fmt_num(pi_item.quantity)}"
                    requests.append(f"수량 PI 기준 {fmt_num(pi_item.quantity)}으로 수정 또는 확인 필요.")
            else:
                qty_diff = f"일치 ({fmt_num(pl_total_qty)})"

        # ── 입수단위 비교 (MAIN만) ─────────────
        if item_class == "MAIN" and pi_item.ea_per_carton is not None and pl_ea is not None:
            if abs(pi_item.ea_per_carton - pl_ea) > 1e-9:
                status = "불일치"
                severity = "높음"
                ea_diff = f"PL {fmt_num(pl_ea)} / PI {fmt_num(pi_item.ea_per_carton)}"
                requests.append(f"입수단위 PI 기준 {fmt_num(pi_item.ea_per_carton)}으로 수정 필요.")
            else:
                ea_diff = f"일치 ({fmt_num(pl_ea)})"

        # ── 카톤수 교차 검증 (MAIN만) ──────────
        pi_ctns_calc: float | None = None
        if item_class == "MAIN" and pi_item.ea_per_carton and pl_total_qty:
            pi_ctns_calc = pl_total_qty / pi_item.ea_per_carton
            if pl_total_ctns and abs(pi_ctns_calc - pl_total_ctns) > 2:
                ctns_diff = f"PL {fmt_num(pl_total_ctns)} CTN / 예상(qty÷ea) {fmt_num(pi_ctns_calc)}"
                if status == "정상":
                    status = "확인필요"
                    severity = "중간"
                    requests.append(f"카톤수 재검토 필요 (예상 {fmt_num(pi_ctns_calc)} CTN).")
            else:
                ctns_diff = f"일치 ({fmt_num(pl_total_ctns)} CTN)"

        results.append(PlCompareRow(
            status=status, severity=severity,
            model=model, description=description, item_class=item_class,
            pi_no=pi_item.pi_no, pi_qty=pi_item.quantity, pl_qty=pl_total_qty,
            qty_diff=qty_diff, pi_ea_per_ctn=pi_item.ea_per_carton, pl_ea_per_ctn=pl_ea,
            ea_diff=ea_diff, pi_ctns_calc=pi_ctns_calc, pl_ctns=pl_total_ctns,
            ctns_diff=ctns_diff, request=" ".join(requests),
            box_qty=box_qty_map.get(model),
            pdf_name=find_doc_name_by_model(pi_docs, model),
        ))

    return results


# ---------------------------------------------------------------------------
# 엑셀 출력 유틸리티
# ---------------------------------------------------------------------------

def autosize(worksheet, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        width = 10
        for cell in column_cells:
            value = norm_text(cell.value)
            if value:
                width = max(width, min(max_width, len(value) + 2))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def style_header(worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def write_rows(worksheet, headers: list[str], rows: Iterable[Iterable[object]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    style_header(worksheet)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(worksheet)


STATUS_FILLS = {
    "정상": PatternFill("solid", fgColor="DCFCE7"),
    "부분선적": PatternFill("solid", fgColor="DBEAFE"),
    "불일치": PatternFill("solid", fgColor="FEE2E2"),
    "PI항목누락": PatternFill("solid", fgColor="FEE2E2"),
    "PI누락": PatternFill("solid", fgColor="FEE2E2"),
    "PI미매칭": PatternFill("solid", fgColor="FEE2E2"),
    "OCR필요": PatternFill("solid", fgColor="FEF3C7"),
    "OCR_FAILED": PatternFill("solid", fgColor="FEE2E2"),
    "PARSED_OCR": PatternFill("solid", fgColor="E0F2FE"),
    "확인필요": PatternFill("solid", fgColor="FEF3C7"),
    "확인": PatternFill("solid", fgColor="FEF3C7"),
    "모델불명": PatternFill("solid", fgColor="F3E8FF"),
}


def apply_status_colors(worksheet) -> None:
    for row in worksheet.iter_rows(min_row=2):
        status = norm_text(row[0].value)
        fill = STATUS_FILLS.get(status)
        if fill:
            for cell in row:
                cell.fill = fill


# ---------------------------------------------------------------------------
# 결과 워크북 작성
# ---------------------------------------------------------------------------

def write_result_workbook(
    output_path: Path,
    input_path: Path,
    excel_path: Path,
    ci_pl_path: Path | None,
    excel_rows: list[ExcelRow],
    pi_docs: list[PiDocument],
    box_compare: list[BoxCompareRow],
    ci_compare: list[CiCompareRow],
    pl_compare: list[PlCompareRow],
    ci_items: list[CiItem],
    pl_items: list[PlItem],
    unsupported_files: list[Path],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. 요약 ─────────────────────────────────
    ws_sum = wb.create_sheet("요약")

    def cnt(rows, status):
        return sum(1 for r in rows if r.status == status)

    all_actions = (
        [r for r in box_compare if r.status != "정상"]
        + [r for r in ci_compare if r.status not in ("정상", "부분선적")]
        + [r for r in pl_compare if r.status not in ("정상", "부분선적")]
    )

    ocr_done = sum(1 for d in pi_docs if d.status == "PARSED_OCR")
    ocr_req = sum(1 for d in pi_docs if d.status == "OCR_REQUIRED")
    ocr_failed = sum(1 for d in pi_docs if d.status == "OCR_FAILED")

    write_rows(ws_sum, ["항목", "값"], [
        ["생성시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["입력", str(input_path)],
        ["박스내용 엑셀", excel_path.name],
        ["CI/PL 파일", ci_pl_path.name if ci_pl_path else "(없음)"],
        ["엑셀 비교 행", len(excel_rows)],
        ["PI PDF 수", len(pi_docs)],
        ["  └ 텍스트 파싱", sum(1 for d in pi_docs if d.status == "PARSED")],
        ["  └ OCR 완료", ocr_done],
        ["  └ OCR 미설치(수동필요)", ocr_req],
        ["  └ OCR 실패", ocr_failed],
        ["CI 항목 수", len(ci_items)],
        ["PL 항목 수", len(pl_items)],
        ["", ""],
        ["[PI vs 박스내용]", ""],
        ["  정상", cnt(box_compare, "정상")],
        ["  불일치", cnt(box_compare, "불일치")],
        ["  PI항목누락", cnt(box_compare, "PI항목누락")],
        ["  PI누락", cnt(box_compare, "PI누락")],
        ["  OCR필요", cnt(box_compare, "OCR필요")],
        ["  확인필요", cnt(box_compare, "확인필요")],
        ["", ""],
        ["[PI vs CI]", ""],
        ["  정상", cnt(ci_compare, "정상")],
        ["  불일치", cnt(ci_compare, "불일치")],
        ["  부분선적", cnt(ci_compare, "부분선적")],
        ["  PI미매칭", cnt(ci_compare, "PI미매칭")],
        ["  OCR필요", cnt(ci_compare, "OCR필요")],
        ["  모델불명", cnt(ci_compare, "모델불명")],
        ["", ""],
        ["[PI vs PL]", ""],
        ["  정상", cnt(pl_compare, "정상")],
        ["  불일치", cnt(pl_compare, "불일치")],
        ["  부분선적", cnt(pl_compare, "부분선적")],
        ["  PI미매칭", cnt(pl_compare, "PI미매칭")],
        ["  OCR필요", cnt(pl_compare, "OCR필요")],
        ["  확인필요", cnt(pl_compare, "확인필요")],
        ["", ""],
        ["수정요청 총계", len(all_actions)],
        ["지원 제외 파일", "\n".join(p.name for p in unsupported_files)],
    ])

    # ── 2. 수정요청 (통합) ──────────────────────
    ws_req = wb.create_sheet("수정요청")
    request_rows = []

    for r in box_compare:
        if r.status != "정상":
            request_rows.append([
                r.status, r.severity, "PI vs 박스내용",
                r.excel.order_no if r.excel else "",
                r.excel.model if r.excel else "",
                r.excel.product if r.excel else "",
                r.diff, r.request, r.pdf_name,
                r.excel.row_no if r.excel else "",
            ])

    for r in ci_compare:
        if r.status not in ("정상", "부분선적"):
            request_rows.append([
                r.status, r.severity, "PI vs CI",
                r.pi_no or "",
                r.model or "",
                r.description,
                r.qty_diff + (" | " + r.price_diff if r.price_diff else ""),
                r.request, r.pdf_name, "",
            ])

    for r in pl_compare:
        if r.status not in ("정상", "부분선적"):
            request_rows.append([
                r.status, r.severity, "PI vs PL",
                r.pi_no or "",
                r.model or "",
                r.description,
                r.qty_diff + (" | " + r.ea_diff if r.ea_diff else ""),
                r.request, r.pdf_name, "",
            ])

    write_rows(ws_req,
        ["상태", "심각도", "비교구분", "주문번호", "모델", "제품/설명", "차이내용", "수정요청", "PDF", "엑셀행"],
        request_rows,
    )
    apply_status_colors(ws_req)

    # ── 3. PI vs 박스내용 ───────────────────────
    ws_box = wb.create_sheet("PI_vs_박스내용")
    write_rows(ws_box, [
        "상태", "심각도", "엑셀행", "주문번호", "모델", "항목구분", "제품",
        "엑셀 구분", "엑셀 수량", "PI 수량", "엑셀 입수", "PI 입수",
        "차이", "수정요청", "PDF", "근거",
    ], (
        [
            r.status, r.severity,
            r.excel.row_no if r.excel else "",
            r.excel.order_no if r.excel else "",
            r.excel.model if r.excel else (r.pi.model if r.pi else ""),
            r.excel.item_class if r.excel else (r.pi.item_class if r.pi else ""),
            r.excel.product if r.excel else "",
            r.excel.description if r.excel else "",
            fmt_num(r.excel.quantity if r.excel else None),
            fmt_num(r.pi.quantity if r.pi else None),
            fmt_num(r.excel.carton_unit if r.excel else None),
            fmt_num(r.pi.ea_per_carton if r.pi else None),
            r.diff, r.request, r.pdf_name, r.evidence,
        ]
        for r in box_compare
    ))
    apply_status_colors(ws_box)

    # ── 4. PI vs CI ─────────────────────────────
    if ci_compare:
        ws_ci = wb.create_sheet("PI_vs_CI")
        write_rows(ws_ci, [
            "상태", "심각도", "모델", "항목구분", "CI 설명",
            "주문번호(PI)", "PI 수량", "CI 수량", "수량 비교",
            "PI 단가", "CI 단가", "단가 비교",
            "PI 총액", "CI 금액", "금액 비교",
            "박스내용 수량(참고)", "수정요청", "PDF",
        ], (
            [
                r.status, r.severity, r.model or "", r.item_class, r.description,
                r.pi_no or "",
                fmt_num(r.pi_qty), fmt_num(r.ci_qty), r.qty_diff,
                fmt_num(r.pi_unit_price), fmt_num(r.ci_unit_price), r.price_diff,
                fmt_num(r.pi_total), fmt_num(r.ci_amount), r.amount_diff,
                fmt_num(r.box_qty), r.request, r.pdf_name,
            ]
            for r in ci_compare
        ))
        apply_status_colors(ws_ci)

    # ── 5. PI vs PL ─────────────────────────────
    if pl_compare:
        ws_pl = wb.create_sheet("PI_vs_PL")
        write_rows(ws_pl, [
            "상태", "심각도", "모델", "항목구분", "PL 설명",
            "주문번호(PI)", "PI 수량", "PL 수량(합계)", "수량 비교",
            "PI 입수", "PL 입수(주)", "입수 비교",
            "PI 카톤(계산)", "PL 카톤(합계)", "카톤 비교",
            "박스내용 수량(참고)", "수정요청", "PDF",
        ], (
            [
                r.status, r.severity, r.model or "", r.item_class, r.description,
                r.pi_no or "",
                fmt_num(r.pi_qty), fmt_num(r.pl_qty), r.qty_diff,
                fmt_num(r.pi_ea_per_ctn), fmt_num(r.pl_ea_per_ctn), r.ea_diff,
                fmt_num(r.pi_ctns_calc), fmt_num(r.pl_ctns), r.ctns_diff,
                fmt_num(r.box_qty), r.request, r.pdf_name,
            ]
            for r in pl_compare
        ))
        apply_status_colors(ws_pl)

    # ── 6. PI 추출 ───────────────────────────────
    ws_pi = wb.create_sheet("PI추출")
    write_rows(ws_pi, [
        "PI번호", "PDF", "페이지", "항목구분", "모델", "제품",
        "입수", "수량", "단가", "금액", "근거",
    ], (
        [
            item.pi_no, item.pdf_name, item.page, item.item_class,
            item.model, item.product,
            fmt_num(item.ea_per_carton), fmt_num(item.quantity),
            fmt_num(item.unit_price), fmt_num(item.total_price),
            item.evidence,
        ]
        for doc in pi_docs
        for item in doc.items
    ))

    # ── 7. CI 추출 ───────────────────────────────
    if ci_items:
        ws_ci_raw = wb.create_sheet("CI추출")
        write_rows(ws_ci_raw, [
            "순번", "모델", "항목구분", "설명", "전압", "수량", "단위", "단가", "금액",
        ], (
            [
                c.line_no, c.model or "", c.item_class, c.description,
                c.volt, fmt_num(c.qty), c.unit,
                fmt_num(c.unit_price), fmt_num(c.amount),
            ]
            for c in ci_items
        ))

    # ── 8. PL 추출 ───────────────────────────────
    if pl_items:
        ws_pl_raw = wb.create_sheet("PL추출")
        write_rows(ws_pl_raw, [
            "순번", "모델", "항목구분", "설명", "CTN범위",
            "입수(EA/CTN)", "카톤수", "수량(PCS)", "순중량", "총중량", "부피(CBM)",
        ], (
            [
                p.line_no, p.model or "", p.item_class, p.description, p.ctn_range,
                fmt_num(p.ea_per_carton), fmt_num(p.ctns), fmt_num(p.qty),
                fmt_num(p.nw), fmt_num(p.gw), fmt_num(p.meas),
            ]
            for p in pl_items
        ))

    # ── 9. 박스내용 원본 ─────────────────────────
    ws_box_raw = wb.create_sheet("박스내용원본")
    write_rows(ws_box_raw, [
        "엑셀행", "주문번호", "로트번호", "모델", "항목구분",
        "제품", "구분", "입수단위", "카톤수", "수량", "비고",
    ], (
        [
            r.row_no, r.order_no, r.lot_no, r.model, r.item_class,
            r.product, r.description,
            fmt_num(r.carton_unit), fmt_num(r.carton_count), fmt_num(r.quantity),
            r.note,
        ]
        for r in excel_rows
    ))

    # ── 10. 파싱 로그 ────────────────────────────
    ws_log = wb.create_sheet("파싱로그")
    write_rows(ws_log, [
        "PDF", "PI번호", "상태", "페이지수", "텍스트문자수",
        "LOT근거", "추출항목수", "오류/메모",
    ], (
        [
            doc.path.name, doc.pi_no, doc.status,
            doc.page_count, doc.char_count, doc.lot_text,
            len(doc.items), doc.error or "",
        ]
        for doc in pi_docs
    ))

    wb.save(output_path)


# ---------------------------------------------------------------------------
# JSON 디버그 출력
# ---------------------------------------------------------------------------

def write_json(
    output_path: Path,
    excel_rows: list[ExcelRow],
    pi_docs: list[PiDocument],
    box_compare: list[BoxCompareRow],
    ci_items: list[CiItem],
    pl_items: list[PlItem],
) -> None:
    payload = {
        "excel_rows": [r.__dict__ for r in excel_rows],
        "pi_docs": [
            {
                "path": str(doc.path),
                "pi_no": doc.pi_no,
                "status": doc.status,
                "page_count": doc.page_count,
                "char_count": doc.char_count,
                "items": [i.__dict__ for i in doc.items],
            }
            for doc in pi_docs
        ],
        "ci_items": [c.__dict__ for c in ci_items],
        "pl_items": [p.__dict__ for p in pl_items],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# 실행 진입점
# ---------------------------------------------------------------------------

def run(input_path: Path, output_dir: Path, keep_json: bool = False) -> Path:
    root, temp_dir = find_input_root(input_path)
    try:
        output_dir.mkdir(parents=True, exist_ok=True)

        print(f"  입력 루트: {root}")

        excel_path = choose_excel(root)
        print(f"  박스내용: {excel_path.name}")

        ci_pl_path = find_ci_pl_file(root)
        print(f"  CI/PL:   {ci_pl_path.name if ci_pl_path else '(없음)'}")

        pdf_paths = find_pdfs(root)
        print(f"  PI PDF:  {len(pdf_paths)}개")

        excel_rows = read_excel_rows(excel_path)
        pi_docs = [read_pdf(path) for path in pdf_paths]

        ci_items: list[CiItem] = []
        pl_items: list[PlItem] = []
        if ci_pl_path:
            ci_items, pl_items = parse_ci_pl(ci_pl_path)
            print(f"  CI 항목: {len(ci_items)}개  /  PL 항목: {len(pl_items)}개")

        # OCR 파싱된 PI 모델코드를 CI/PL/박스내용 기준으로 보정
        known_models: frozenset[str] = frozenset(
            m for m in (
                [c.model for c in ci_items if c.model]
                + [p.model for p in pl_items if p.model]
                + [r.model for r in excel_rows if r.model]
            )
        )
        if any(d.status == "PARSED_OCR" for d in pi_docs):
            normalize_ocr_models(pi_docs, known_models)
            ocr_count = sum(1 for d in pi_docs if d.status == "PARSED_OCR")
            print(f"  OCR 완료: {ocr_count}개 PDF (모델코드 자동보정 적용)")

        box_compare = compare_pi_vs_box(excel_rows, pi_docs)
        ci_compare = compare_pi_vs_ci(pi_docs, ci_items, excel_rows)
        pl_compare = compare_pi_vs_pl(pi_docs, pl_items, excel_rows)

        unsupported = sorted(
            p for p in root.rglob("*.xls")
            if ci_pl_path is None or p != ci_pl_path
        )

        output_path = output_dir / "pi_ci_pl_review.xlsx"
        write_result_workbook(
            output_path=output_path,
            input_path=input_path,
            excel_path=excel_path,
            ci_pl_path=ci_pl_path,
            excel_rows=excel_rows,
            pi_docs=pi_docs,
            box_compare=box_compare,
            ci_compare=ci_compare,
            pl_compare=pl_compare,
            ci_items=ci_items,
            pl_items=pl_items,
            unsupported_files=unsupported,
        )

        if keep_json:
            write_json(
                output_dir / "pi_ci_pl_debug.json",
                excel_rows, pi_docs, box_compare, ci_items, pl_items,
            )

        return output_path
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="PI / CI / PL을 PI 기준으로 비교하여 검토 결과 엑셀을 생성합니다."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default="/Users/noroovirus/Downloads/물류.zip",
        help="입력 ZIP 또는 압축 해제 폴더 경로",
    )
    parser.add_argument(
        "-o", "--output-dir",
        default="/Users/noroovirus/Downloads/new",
        help="결과 엑셀을 저장할 폴더",
    )
    parser.add_argument(
        "--debug-json", action="store_true",
        help="디버그용 JSON도 함께 저장합니다.",
    )
    args = parser.parse_args()

    output_path = run(Path(args.input), Path(args.output_dir), keep_json=args.debug_json)
    print(f"\n✅ 완료: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
