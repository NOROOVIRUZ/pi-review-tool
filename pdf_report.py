"""민트 토스 스타일 PDF 보고서 — Pretendard 폰트, 캔버스 색 원 아이콘"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Flowable,
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ---------------------------------------------------------------------------
# 폰트 등록
# ---------------------------------------------------------------------------
_FONT_DIR = Path(__file__).parent / "static" / "fonts"

pdfmetrics.registerFont(TTFont("PD",  _FONT_DIR / "Pretendard-Regular.ttf"))
pdfmetrics.registerFont(TTFont("PDB", _FONT_DIR / "Pretendard-Bold.ttf"))
pdfmetrics.registerFont(TTFont("PDS", _FONT_DIR / "Pretendard-SemiBold.ttf"))
pdfmetrics.registerFontFamily("Pretendard", normal="PD", bold="PDB")

KR  = "PD"
KRB = "PDB"
KRS = "PDS"

# ---------------------------------------------------------------------------
# 민트 컬러 팔레트
# ---------------------------------------------------------------------------
MINT      = colors.HexColor("#00BFA5")   # 메인 민트
MINT_DARK = colors.HexColor("#00897B")   # 헤더 다크 민트
MINT_MID  = colors.HexColor("#4DB6AC")   # 중간 민트
MINT_PALE = colors.HexColor("#E0F7FA")   # 아주 연한 민트 배경
MINT_LINE = colors.HexColor("#B2DFDB")   # 구분선
TEXT_DARK = colors.HexColor("#1B1C1E")   # 본문 어두운
TEXT_GRAY = colors.HexColor("#6B7280")   # 서브 텍스트
TEXT_LIGHT= colors.HexColor("#9CA3AF")   # 힌트
WHITE     = colors.white
BG_CARD   = colors.HexColor("#F9FAFB")   # 카드 배경
BORDER    = colors.HexColor("#E5E7EB")

# 상태 색상
S_COLOR = {
    "정상":       colors.HexColor("#10B981"),
    "불일치":     colors.HexColor("#EF4444"),
    "부분선적":   colors.HexColor("#3B82F6"),
    "PI미매칭":   colors.HexColor("#F59E0B"),
    "PI항목누락": colors.HexColor("#EF4444"),
    "PI누락":     colors.HexColor("#EF4444"),
    "OCR필요":    colors.HexColor("#EAB308"),
    "확인필요":   colors.HexColor("#F59E0B"),
    "모델불명":   colors.HexColor("#8B5CF6"),
    "PARSED_OCR": colors.HexColor("#0EA5E9"),
}
S_BG = {
    "정상":       colors.HexColor("#DCFCE7"),
    "불일치":     colors.HexColor("#FEE2E2"),
    "부분선적":   colors.HexColor("#DBEAFE"),
    "PI미매칭":   colors.HexColor("#FEF3C7"),
    "PI항목누락": colors.HexColor("#FEE2E2"),
    "PI누락":     colors.HexColor("#FEE2E2"),
    "OCR필요":    colors.HexColor("#FEF9C3"),
    "확인필요":   colors.HexColor("#FEF3C7"),
    "모델불명":   colors.HexColor("#F3E8FF"),
    "PARSED_OCR": colors.HexColor("#E0F2FE"),
}

PAGE_W, PAGE_H = A4
MARGIN = 18 * mm
INNER_W = PAGE_W - 2 * MARGIN


# ---------------------------------------------------------------------------
# 스타일 헬퍼
# ---------------------------------------------------------------------------
def _sty(name: str, **kw) -> ParagraphStyle:
    base = getSampleStyleSheet()["Normal"]
    kw.setdefault("fontName", KR)
    return ParagraphStyle(name, parent=base, **kw)


S_BODY  = _sty("Body",  fontSize=9,  textColor=TEXT_DARK, leading=13)
S_SMALL = _sty("Small", fontSize=8,  textColor=TEXT_GRAY, leading=11)
S_TINY  = _sty("Tiny",  fontSize=7,  textColor=TEXT_LIGHT, leading=10)
S_NUM   = _sty("Num",   fontSize=18, textColor=TEXT_DARK, leading=22, fontName=KRB, alignment=TA_CENTER)
S_BADGE = _sty("Badge", fontSize=8,  leading=11, alignment=TA_CENTER)
S_CTR   = _sty("Ctr",   fontSize=8,  textColor=TEXT_DARK, leading=11, alignment=TA_CENTER)
S_FOOT  = _sty("Foot",  fontSize=7,  textColor=TEXT_LIGHT, leading=10, alignment=TA_CENTER)
S_TH    = _sty("TH",    fontSize=8,  textColor=WHITE, leading=11, alignment=TA_CENTER, fontName=KRB)
S_TD    = _sty("TD",    fontSize=8,  textColor=TEXT_DARK, leading=11)
S_TDCTR = _sty("TDCTR", fontSize=8,  textColor=TEXT_DARK, leading=11, alignment=TA_CENTER)


def _p(text, style) -> Paragraph:
    return Paragraph(str(text) if text is not None else "", style)


# ---------------------------------------------------------------------------
# 색 원 아이콘 Flowable (이모지 대체)
# ---------------------------------------------------------------------------
class ColorDot(Flowable):
    """인라인 색 원 — 상태 표시 아이콘용"""
    def __init__(self, color, radius=4, label="", font=KR, fontsize=8):
        super().__init__()
        self.color = color
        self.r = radius
        self.label = label
        self.font = font
        self.fontsize = fontsize
        self.width = radius * 2 + (len(label) * fontsize * 0.6 if label else 0) + 6
        self.height = max(radius * 2, fontsize + 2)

    def draw(self):
        c = self.canv
        c.setFillColor(self.color)
        c.circle(self.r, self.height / 2, self.r, fill=1, stroke=0)
        if self.label:
            c.setFillColor(TEXT_DARK)
            c.setFont(self.font, self.fontsize)
            c.drawString(self.r * 2 + 4, (self.height - self.fontsize) / 2, self.label)


# ---------------------------------------------------------------------------
# 섹션 헤더 (민트 좌측 바 + 굵은 제목)
# ---------------------------------------------------------------------------
class SectionHeader(Flowable):
    def __init__(self, symbol: str, title: str, color=MINT):
        super().__init__()
        self.symbol = symbol
        self.title = title
        self.color = color
        self.width = INNER_W
        self.height = 28

    def draw(self):
        c = self.canv
        # 좌측 민트 바
        c.setFillColor(self.color)
        c.rect(0, 2, 4, self.height - 4, fill=1, stroke=0)
        # 배경
        c.setFillColor(MINT_PALE)
        c.rect(4, 0, self.width - 4, self.height, fill=1, stroke=0)
        # 심볼
        c.setFillColor(self.color)
        c.setFont(KRB, 13)
        c.drawString(12, 8, self.symbol)
        sym_w = c.stringWidth(self.symbol, KRB, 13)
        # 제목
        c.setFillColor(TEXT_DARK)
        c.setFont(KRB, 12)
        c.drawString(12 + sym_w + 6, 8, self.title)


# ---------------------------------------------------------------------------
# 페이지 헤더/푸터 (canvas 직접 그리기)
# ---------------------------------------------------------------------------
def _page_one(canvas, doc):
    canvas.saveState()
    # ── 민트 그라디언트 헤더 ──
    # 다크 민트 띠
    canvas.setFillColor(MINT_DARK)
    canvas.rect(0, PAGE_H - 55 * mm, PAGE_W, 55 * mm, fill=1, stroke=0)
    # 연한 민트 사선 장식 (오른쪽)
    canvas.setFillColor(MINT_MID)
    canvas.setStrokeAlpha(0)
    p = canvas.beginPath()
    p.moveTo(PAGE_W - 60 * mm, PAGE_H)
    p.lineTo(PAGE_W, PAGE_H)
    p.lineTo(PAGE_W, PAGE_H - 55 * mm)
    p.lineTo(PAGE_W - 90 * mm, PAGE_H - 55 * mm)
    p.close()
    canvas.drawPath(p, fill=1, stroke=0)

    # ── 텍스트 ──
    canvas.setFillColor(WHITE)
    canvas.setFont(KRB, 20)
    canvas.drawString(MARGIN, PAGE_H - 22 * mm, "PI / CI / PL")
    canvas.setFont(KRB, 14)
    canvas.drawString(MARGIN, PAGE_H - 32 * mm, "물류 검토 보고서")
    canvas.setFont(KR, 9)
    canvas.setFillColor(colors.HexColor("#A7F3D0"))
    canvas.drawString(MARGIN, PAGE_H - 41 * mm, "Made by noroovirus  ·  PI · CI · PL 3-way 자동 비교 분석")

    # noroovirus 태그 (우측 상단)
    tag_x = PAGE_W - MARGIN - 60 * mm
    canvas.setFillColor(colors.HexColor("#CCFBF1"))
    canvas.roundRect(tag_x, PAGE_H - 28 * mm, 55 * mm, 9 * mm, 4, fill=1, stroke=0)
    canvas.setFillColor(MINT_DARK)
    canvas.setFont(KRB, 8)
    canvas.drawCentredString(tag_x + 27.5 * mm, PAGE_H - 23 * mm, "★  noroovirus")

    # ── 페이지 번호 ──
    canvas.setFillColor(TEXT_LIGHT)
    canvas.setFont(KR, 7)
    canvas.drawCentredString(PAGE_W / 2, 8 * mm, f"— {doc.page} —")
    canvas.restoreState()


def _page_sub(canvas, doc):
    canvas.saveState()
    # 얇은 민트 헤더 바
    canvas.setFillColor(MINT)
    canvas.rect(0, PAGE_H - 10 * mm, PAGE_W, 10 * mm, fill=1, stroke=0)
    canvas.setFillColor(WHITE)
    canvas.setFont(KR, 7.5)
    canvas.drawString(MARGIN, PAGE_H - 7 * mm, "PI / CI / PL 물류 검토 보고서  |  Made by noroovirus")
    canvas.setFont(KR, 7.5)
    canvas.drawRightString(PAGE_W - MARGIN, PAGE_H - 7 * mm, "자동 생성 문서")
    # 하단 민트 라인
    canvas.setStrokeColor(MINT_LINE)
    canvas.setLineWidth(0.5)
    canvas.line(MARGIN, 14 * mm, PAGE_W - MARGIN, 14 * mm)
    canvas.setFillColor(TEXT_LIGHT)
    canvas.setFont(KR, 7)
    canvas.drawCentredString(PAGE_W / 2, 9 * mm, f"— {doc.page} —")
    canvas.restoreState()


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------
def _read_sheet(wb, name):
    if name not in wb.sheetnames:
        return []
    return [[c if c is not None else "" for c in row]
            for row in wb[name].iter_rows(values_only=True)]


def _hr():
    return HRFlowable(width="100%", thickness=0.5, color=MINT_LINE,
                      spaceAfter=5, spaceBefore=5)


# ---------------------------------------------------------------------------
# 상태 뱃지 셀 (colored bg + symbol + text)
# ---------------------------------------------------------------------------
def _status_cell(status: str) -> Paragraph:
    fg  = S_COLOR.get(status, TEXT_GRAY)
    sym = {"정상": "✓", "불일치": "✗", "부분선적": "▶",
           "PI미매칭": "▲", "PI항목누락": "✗", "PI누락": "✗",
           "OCR필요": "◎", "확인필요": "▲", "모델불명": "◆"}.get(status, "●")
    fg_hex = fg.hexval()
    style = _sty(f"sb_{status}", fontSize=8, leading=11, alignment=TA_CENTER,
                 textColor=fg,
                 backColor=S_BG.get(status, BG_CARD))
    return Paragraph(f"{sym} {status}", style)


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
def generate(excel_path: Path, output_dir: Path) -> Path:
    wb = openpyxl.load_workbook(excel_path)
    sum_rows  = _read_sheet(wb, "요약")
    req_rows  = _read_sheet(wb, "수정요청")

    # 요약 dict (key stripped)
    summary: dict[str, str] = {}
    for row in sum_rows[1:]:
        if row and row[0] != "":
            summary[str(row[0]).strip()] = str(row[1]).strip() if row[1] != "" else ""

    pdf_path = output_dir / "pi_ci_pl_report.pdf"

    doc = SimpleDocTemplate(
        str(pdf_path), pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=62 * mm, bottomMargin=20 * mm,
    )

    story: list = []

    # ══════════════════════════════════════════════════
    # 1. 파일 정보
    # ══════════════════════════════════════════════════
    story.append(SectionHeader("●", "검토 파일 정보"))
    story.append(Spacer(1, 4 * mm))

    ocr_done = summary.get("└ OCR 완료", summary.get("  └ OCR 완료", "0"))
    txt_done = summary.get("└ 텍스트 파싱", summary.get("  └ 텍스트 파싱", "0"))
    pi_count = summary.get("PI PDF 수", "?")

    info = [
        ["생성 시각",  summary.get("생성시각", "-")],
        ["입력 파일",  Path(summary.get("입력", "-")).name],
        ["박스내용",   summary.get("박스내용 엑셀", "-")],
        ["CI / PL",    summary.get("CI/PL 파일", "-")],
        ["PI PDF",     f"{pi_count}개  ·  텍스트: {txt_done}  ·  OCR 완료: {ocr_done}"],
        ["CI 항목",    summary.get("CI 항목 수", "-") + "개"],
        ["PL 항목",    summary.get("PL 항목 수", "-") + "개"],
    ]

    info_tbl = Table(
        [[_p(k, _sty(f"ik{i}", fontSize=8, textColor=MINT_DARK, fontName=KRS, leading=11)),
          _p(v, S_SMALL)] for i, (k, v) in enumerate(info)],
        colWidths=[28 * mm, INNER_W - 28 * mm],
    )
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), MINT_PALE),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [WHITE, BG_CARD]),
        ("GRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("FONTNAME", (0, 0), (-1, -1), KR),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 8 * mm))

    # ══════════════════════════════════════════════════
    # 2. 비교 결과 요약 카드 3개
    # ══════════════════════════════════════════════════
    story.append(SectionHeader("◆", "비교 결과 요약"))
    story.append(Spacer(1, 4 * mm))

    STAT_DEFS = [
        ("PI vs 박스내용", [
            ("정상", "✓"), ("불일치", "✗"), ("부분선적", "▶"),
            ("PI항목누락", "✗"), ("PI누락", "✗"), ("OCR필요", "◎"), ("확인필요", "▲"),
        ]),
        ("PI vs CI", [
            ("정상", "✓"), ("불일치", "✗"), ("부분선적", "▶"),
            ("PI미매칭", "▲"), ("OCR필요", "◎"), ("모델불명", "◆"),
        ]),
        ("PI vs PL", [
            ("정상", "✓"), ("불일치", "✗"), ("부분선적", "▶"),
            ("PI미매칭", "▲"), ("OCR필요", "◎"), ("확인필요", "▲"),
        ]),
    ]

    def _stat_card(label: str, items: list) -> Table:
        # 헤더 행
        hdr = Table([[_p(label, _sty(f"sh_{label}", fontSize=8.5, textColor=WHITE,
                                     fontName=KRB, leading=12, alignment=TA_CENTER))]],
                    colWidths=[INNER_W / 3 - 4 * mm])
        hdr.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), MINT),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0,0), (-1, -1), 5),
        ]))

        rows = []
        for status, sym in items:
            val = int(summary.get(status, 0) or 0)
            fg  = S_COLOR.get(status, TEXT_GRAY)
            bg  = S_BG.get(status, WHITE) if val > 0 else WHITE
            num_style = _sty(f"n_{status}", fontSize=13, textColor=fg if val > 0 else TEXT_LIGHT,
                             fontName=KRB, leading=16, alignment=TA_CENTER)
            lbl_style = _sty(f"l_{status}", fontSize=7.5, textColor=fg if val > 0 else TEXT_LIGHT,
                             leading=10, alignment=TA_CENTER)
            cell = Table(
                [[_p(sym + " " + status, lbl_style)],
                 [_p(str(val), num_style)]],
                colWidths=[INNER_W / 3 - 4 * mm],
            )
            cell.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg),
                ("LEFTPADDING",  (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING",   (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
                ("LINEBELOW", (0, 0), (-1, -1), 0.3, BORDER),
            ]))
            rows.append(cell)

        outer = Table(
            [[hdr]] + [[r] for r in rows],
            colWidths=[INNER_W / 3 - 4 * mm],
        )
        outer.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, MINT_LINE),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING",   (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
        ]))
        return outer

    cards = [_stat_card(lbl, items) for lbl, items in STAT_DEFS]
    card_gap = 2 * mm
    cards_tbl = Table(
        [cards],
        colWidths=[INNER_W / 3 - card_gap] * 3,
        hAlign="CENTER",
    )
    cards_tbl.setStyle(TableStyle([
        ("LEFTPADDING",  (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(cards_tbl)
    story.append(Spacer(1, 5 * mm))

    # 수정요청 총계 배너
    total = summary.get("수정요청 총계", "0")
    total_s = _sty("ttl", fontSize=12, textColor=WHITE, fontName=KRB, leading=16)
    total_n = _sty("tnum", fontSize=20, textColor=colors.HexColor("#CCFBF1"),
                   fontName=KRB, leading=24, alignment=TA_RIGHT)

    banner = Table(
        [[_p("▶  수정요청 총계", total_s), _p(str(total) + "건", total_n)]],
        colWidths=[None, 28 * mm],
    )
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), MINT_DARK),
        ("LEFTPADDING",  (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING",   (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(banner)

    # ══════════════════════════════════════════════════
    # 3. 수정요청 상세 테이블
    # ══════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Spacer(1, 3 * mm))
    story.append(SectionHeader("▲", "수정요청 상세 목록", color=colors.HexColor("#EF4444")))
    story.append(Spacer(1, 4 * mm))

    if len(req_rows) > 1:
        # 상태 | 비교구분 | 주문번호 | 모델 (+ PI파일명) | 제품/설명 | 차이내용/수정요청
        COL_W = [18*mm, 16*mm, 22*mm, 32*mm, 30*mm, 49*mm]
        HEADERS = ["상태", "비교구분", "주문번호", "모델 / PI 파일", "제품/설명", "차이내용 / 수정요청"]

        tdata = [[_p(h, S_TH) for h in HEADERS]]

        for row in req_rows[1:]:
            status   = str(row[0]) if row[0] != "" else ""
            compare  = str(row[2]).replace("PI vs ", "") if len(row) > 2 else ""
            pi_no    = str(row[3]) if len(row) > 3 else ""
            model    = str(row[4]) if len(row) > 4 else ""
            product  = str(row[5]) if len(row) > 5 else ""
            diff     = str(row[6]) if len(row) > 6 else ""
            req      = str(row[7]) if len(row) > 7 else ""
            pdf_name = str(row[8]) if len(row) > 8 else ""

            # PI 파일명 정리: 확장자 제거 + 최대 26자
            pdf_short = pdf_name.removesuffix(".pdf")
            if len(pdf_short) > 26:
                pdf_short = pdf_short[:23] + "..."

            combined = diff if not req or req == diff else f"{diff}\n→ {req}"

            fg = S_COLOR.get(status, TEXT_GRAY)
            sym = {"정상":"✓","불일치":"✗","부분선적":"▶","PI미매칭":"▲",
                   "PI항목누락":"✗","PI누락":"✗","OCR필요":"◎",
                   "확인필요":"▲","모델불명":"◆"}.get(status, "●")
            badge_s = _sty(f"bs_{status}", fontSize=7.5, textColor=fg, fontName=KRS,
                           leading=11, alignment=TA_CENTER,
                           backColor=S_BG.get(status, BG_CARD))

            # 모델 + PI파일명 2줄 셀
            model_cell = _p(
                f'<font name="{KRB}" size="8">{model}</font>'
                + (f'<br/><font name="{KR}" size="6" color="#9CA3AF">{pdf_short}</font>' if pdf_short else ""),
                _sty(f"mc_{model}", fontSize=8, leading=11),
            )

            tdata.append([
                _p(f"{sym} {status}", badge_s),
                _p(compare,  S_TDCTR),
                _p(pi_no,    S_TDCTR),
                model_cell,
                _p(product,  _sty("prod", fontSize=7.5, textColor=TEXT_GRAY, leading=10)),
                _p(combined, S_TD),
            ])

        req_tbl = Table(tdata, colWidths=COL_W, repeatRows=1)
        tbl_style = [
            # 헤더
            ("BACKGROUND",   (0, 0), (-1, 0), MINT_DARK),
            ("FONTNAME",     (0, 0), (-1, 0), KRB),
            ("FONTSIZE",     (0, 0), (-1, 0), 8),
            ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
            ("TEXTCOLOR",    (0, 0), (-1, 0), WHITE),
            # 전체 셀
            ("FONTNAME",     (0, 1), (-1, -1), KR),
            ("FONTSIZE",     (0, 1), (-1, -1), 8),
            ("GRID",         (0, 0), (-1, -1), 0.3, BORDER),
            ("LEFTPADDING",  (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING",   (0, 1), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 1), (-1, -1), 5),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ]
        # 홀짝 행 배경 + 상태 배경
        for i, row in enumerate(req_rows[1:], start=1):
            status = str(row[0]) if row[0] != "" else ""
            bg = S_BG.get(status, BG_CARD if i % 2 == 0 else WHITE)
            tbl_style.append(("BACKGROUND", (0, i), (-1, i), bg))
            # 왼쪽 테두리 색 강조
            fg = S_COLOR.get(status, MINT_LINE)
            tbl_style.append(("LINEAFTER", (-1, i), (-1, i), 0, WHITE))

        req_tbl.setStyle(TableStyle(tbl_style))
        story.append(req_tbl)
    else:
        story.append(_p("✓  수정요청 항목이 없습니다. 모든 항목 정상.", S_BODY))

    story.append(Spacer(1, 8 * mm))
    story.append(_p("이 보고서는 PI / CI / PL 검토 툴에 의해 자동 생성되었습니다.  |  Made by noroovirus", S_FOOT))

    doc.build(story, onFirstPage=_page_one, onLaterPages=_page_sub)
    return pdf_path
